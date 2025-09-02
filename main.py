# main.py
# usage:
#   python main.py [info.txt] [input.xlsx] [output.xlsx]
import re
import sys
from pathlib import Path
from dataclasses import dataclass
from typing import List, Dict, Tuple, Optional

from openpyxl import load_workbook
from openpyxl.styles import PatternFill

SHEET_ALIASES = {
    "BW": ["brickwork"],
    "SR": ["sideraise", "siderise", "side raise", "side-raise", "side rise"],
}

# ===== Дані рядка з info.txt =====
@dataclass
class Entry:
    block: str
    level: int
    side: str     # East/West/North/South
    qnum: int
    count: int
    kind: str     # "BW" або "SR"

# ===== Кольори =====
def color_fill_for_count(n: int) -> PatternFill:
    if n <= 6:
        return PatternFill(patternType="solid", fgColor="FFFF0000")  # red
    if 7 <= n <= 20:
        return PatternFill(patternType="solid", fgColor="FFFFFF00")  # yellow
    return PatternFill(patternType="solid", fgColor="FF00B050")      # green

# ===== Парсинг info.txt =====
LINE_RE = re.compile(
    r"""
    ^\s*(?P<count>\d+)\s+pictures\s*-\s*
    Block\s+(?P<block>[A-Z])\s+
    L(?P<level>\d{1,2})\s+
    (?P<side>North|South|East|West)\s+
    (?P<kind>BW|SR)\b
    .*?/\s*(?P<qnum>\d+)\b
    """,
    re.IGNORECASE | re.VERBOSE,
)

def parse_txt(txt_path: Path) -> List[Entry]:
    out: List[Entry] = []
    with txt_path.open("r", encoding="utf-8") as f:
        for raw in f:
            m = LINE_RE.search(raw)
            if not m:
                continue
            out.append(
                Entry(
                    block=m.group("block").upper(),
                    level=int(m.group("level")),
                    side=m.group("side").capitalize(),
                    qnum=int(m.group("qnum")),
                    count=int(m.group("count")),
                    kind=m.group("kind").upper(),
                )
            )
    return out

# ===== Витяг локації з колонки A (строго: має бути сторона) =====
LOC_STRICT_RE = re.compile(
    r"""
    Block\s+(?P<block>[A-Z])\b .*?
    (?:
        (?P<gl>\b0{0,2}\s*-\s*Ground\s+Level\b) |
        (?:\b\d{1,2}\s*-\s*Level\s+(?P<lvl2>\d{1,2})\b) |
        (?:\bLevel\s+(?P<lvl1>\d{1,2})\b) |
        (?:\bL(?P<lvl0>\d{1,2})\b)
    )
    .*/\s*(?P<side>East|West|North|South)\s+Elevation\b
    """,
    re.IGNORECASE | re.VERBOSE,
)

def parse_loc(text: str) -> Optional[Tuple[str, int, str]]:
    if not isinstance(text, str):
        return None
    m = LOC_STRICT_RE.search(text)
    if not m:
        return None
    block = m.group("block").upper()
    level = 0 if m.group("gl") else next(int(g) for g in (m.group("lvl2"), m.group("lvl1"), m.group("lvl0")) if g)
    side = m.group("side").capitalize()
    return (block, level, side)

# ===== Підбір аркуша під тип (BW/SR) =====
def pick_sheet_for_kind(wb, kind: str):
    aliases = [x.lower() for x in SHEET_ALIASES.get(kind, [])]
    for ws in wb.worksheets:
        title = ws.title.lower()
        if any(a in title for a in aliases):
            return ws
    return None

# ===== Індекси по аркушу =====
def build_q_header_map(ws, header_row=1) -> Dict[int, int]:
    """
    Повертає {номер питання -> індекс колонки}.
    Підтримує заголовки як з пробілом, так і без: "4 All..." та "4All..."
    """
    qmap: Dict[int, int] = {}
    for j, cell in enumerate(ws[header_row], start=1):
        txt = cell.value
        if isinstance(txt, str):
            m = re.match(r"^\s*(\d+)", txt)   # БЕЗ \b!
            if m:
                qmap[int(m.group(1))] = j
    return qmap

def build_row_map(ws, header_row=1) -> Dict[Tuple[str, int, str], int]:
    """
    Повертає {(Block, Level, Side) -> row_index} для рядків, де A — повна локація.
    """
    rmap: Dict[Tuple[str, int, str], int] = {}
    for r in range(header_row + 1, ws.max_row + 1):
        loc = parse_loc(ws.cell(row=r, column=1).value)
        if loc and loc not in rmap:
            rmap[loc] = r
    return rmap

def clear_non_location_rows(ws, qmap: Dict[int, int], header_row=1):
    """
    Очищає всі клітинки з питаннями у рядках, де A НЕ є повною локацією.
    Це прибирає «старі» помилкові цифри (як у рядку 2).
    """
    cols = list(qmap.values())
    if not cols:
        return
    for r in range(header_row + 1, ws.max_row + 1):
        if parse_loc(ws.cell(row=r, column=1).value) is None:
            for c in cols:
                cell = ws.cell(row=r, column=c)
                cell.value = None
                cell.fill = PatternFill()  # зняти заливку

# ===== Запис одиничного значення =====
def write_one(ws, r: int, c: int, value: int):
    cell = ws.cell(row=r, column=c)
    cell.value = value
    cell.fill = color_fill_for_count(value)

# ===== Основний процес =====
def process(entries: List[Entry], xlsx_path: Path, out_path: Path):
    wb = load_workbook(str(xlsx_path))

    sheet_cache = {}
    header_cache = {}
    row_cache = {}
    cleared = set()

    written = 0
    skipped = []

    for e in entries:
        # 1) вибрати аркуш за типом
        ws = sheet_cache.get(e.kind)
        if ws is None:
            ws = pick_sheet_for_kind(wb, e.kind)
            sheet_cache[e.kind] = ws
            if ws is None:
                skipped.append((e, f"sheet for kind '{e.kind}' not found"))
                continue

        # 2) побудувати індекси, одноразово
        if ws not in header_cache:
            header_cache[ws] = build_q_header_map(ws, header_row=1)
        if ws not in row_cache:
            row_cache[ws] = build_row_map(ws, header_row=1)
        # 3) очистити «не-локації» один раз на аркуш
        if ws not in cleared:
            clear_non_location_rows(ws, header_cache[ws], header_row=1)
            cleared.add(ws)

        qmap = header_cache[ws]
        rmap = row_cache[ws]

        row = rmap.get((e.block, e.level, e.side))
        col = qmap.get(e.qnum)

        if row is None:
            skipped.append((e, "location not found in column A"))
            continue
        if col is None:
            skipped.append((e, "question number not found in header"))
            continue

        write_one(ws, row, col, e.count)
        written += 1

    wb.save(out_path)

    print(f"✅ Done. Written: {written}")
    if skipped:
        print("ℹ️ Skipped:")
        for e, why in skipped:
            print(f"  - {e.kind} | Block {e.block} L{e.level} {e.side} / Q{e.qnum}: {why}")
    print(f"📄 Saved to: {out_path.resolve()}")

# ===== CLI =====
def main():
    info_arg = sys.argv[1] if len(sys.argv) > 1 else "info.txt"
    in_arg   = sys.argv[2] if len(sys.argv) > 2 else "output.xlsx"
    out_arg  = sys.argv[3] if len(sys.argv) > 3 else "output_filled.xlsx"

    txt_path  = Path(info_arg).expanduser()
    xlsx_path = Path(in_arg).expanduser()
    out_path  = Path(out_arg).expanduser()

    assert txt_path.exists(),  f"Не знайдено {info_arg} (шлях: {txt_path})"
    assert xlsx_path.exists(), f"Не знайдено {in_arg} (шлях: {xlsx_path})"

    entries = parse_txt(txt_path)
    process(entries, xlsx_path, out_path)

if __name__ == "__main__":
    main()
